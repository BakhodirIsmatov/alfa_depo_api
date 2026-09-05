from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from html import escape
from io import BytesIO
from pathlib import Path
from typing import Any

import reportlab
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.schemas.report import (
    DailyStockReportResponse,
    ProductReportFilters,
    ProductReportItem,
    ProductReportSummary,
    ReportFormat,
    ReportLanguage,
)

A4_LANDSCAPE_WIDTH_MM = 297
PNG_DPI = 300
MM_PER_INCH = 25.4
PNG_MAX_WIDTH = int(A4_LANDSCAPE_WIDTH_MM / MM_PER_INCH * PNG_DPI)
TABLE_MIN_WIDTH_RATIO = 0.82
PNG_HEADER_MIN_HEIGHT = 176
PNG_HEADER_HORIZONTAL_PADDING = 40
PNG_HEADER_VERTICAL_PADDING = 28
PNG_HEADER_TEXT_GAP = 12
PNG_SUMMARY_TOP_PADDING = 36
PNG_SUMMARY_BOTTOM_PADDING = 32
PNG_TABLE_HEADER_MIN_HEIGHT = 76
PNG_TABLE_ROW_MIN_HEIGHT = 60
PNG_CELL_HORIZONTAL_PADDING = 18
PNG_CELL_VERTICAL_PADDING = 12
PNG_MIN_COLUMN_WIDTH = 96
PNG_MAX_UNBROKEN_WIDTH = 240
PNG_TITLE_LINE_HEIGHT = 72
PNG_SUBTITLE_LINE_HEIGHT = 40
PNG_TABLE_HEADER_LINE_HEIGHT = 38
PNG_TABLE_BODY_LINE_HEIGHT = 38
PNG_ACCENT_HEIGHT = 6
PNG_GRID_LINE_WIDTH = 2


@dataclass(frozen=True)
class ExportColumn:
    key: str
    title: str
    width: int  # Relative growth weight and spreadsheet soft-width baseline.


@dataclass(frozen=True)
class ReportDocument:
    title: str
    subtitle: str
    columns: list[ExportColumn]
    rows: list[dict[str, Any]]
    summary: list[tuple[str, Any]]
    filename: str


@dataclass(frozen=True)
class ExportArtifact:
    content: bytes
    media_type: str
    filename: str


TRANSLATIONS = {
    ReportLanguage.TURKISH: {
        "product_report": "Ürün Raporu",
        "daily_report": "Günlük Stok Hareket Raporu",
        "generated": "Oluşturulma",
        "code": "Ürün Kodu",
        "name": "Ürün Adı",
        "brand": "Marka",
        "color": "Renk",
        "lot": "Lot",
        "current": "Mevcut (kg)",
        "count": "Ürün sayısı",
        "status": "Durum",
        "created": "Oluşturuldu",
        "total_products": "Toplam ürün",
        "total_stock": "Toplam stok (kg)",
        "low": "Düşük stok",
        "out": "Stokta yok",
        "date": "Tarih",
        "date_range": "Tarih aralığı",
        "opening": "Başlangıç (kg)",
        "closing": "Kapanış (kg)",
        "stock_in": "Giriş (kg)",
        "stock_out": "Çıkış (kg)",
        "adjustment_in": "Pozitif düzeltme",
        "adjustment_out": "Negatif düzeltme",
        "net": "Net değişim",
        "transactions": "İşlem",
        "affected": "Etkilenen ürün",
        "normal": "Normal",
        "movement_type": "Hareket türü",
        "movement_all": "Tümü",
        "movement_in": "Giriş",
        "movement_out": "Çıkış",
    },
    ReportLanguage.ENGLISH: {
        "product_report": "Product Report",
        "daily_report": "Daily Stock Movement Report",
        "generated": "Generated",
        "code": "Product Code",
        "name": "Product Name",
        "brand": "Brand",
        "color": "Color",
        "lot": "Lot",
        "current": "Current (kg)",
        "count": "Product count",
        "status": "Status",
        "created": "Created",
        "total_products": "Total products",
        "total_stock": "Total stock (kg)",
        "low": "Low stock",
        "out": "Out of stock",
        "date": "Date",
        "date_range": "Date range",
        "opening": "Opening (kg)",
        "closing": "Closing (kg)",
        "stock_in": "Stock in (kg)",
        "stock_out": "Stock out (kg)",
        "adjustment_in": "Positive adjustment",
        "adjustment_out": "Negative adjustment",
        "net": "Net change",
        "transactions": "Transactions",
        "affected": "Affected products",
        "normal": "Normal",
        "movement_type": "Movement type",
        "movement_all": "All",
        "movement_in": "Stock in",
        "movement_out": "Stock out",
    },
    ReportLanguage.UZBEK: {
        "product_report": "Mahsulotlar hisoboti",
        "daily_report": "Kunlik ombor harakatlari hisoboti",
        "generated": "Yaratilgan vaqt",
        "code": "Mahsulot kodi",
        "name": "Mahsulot nomi",
        "brand": "Marka",
        "color": "Rang",
        "lot": "Lot",
        "current": "Mavjud (kg)",
        "count": "Mahsulot soni",
        "status": "Holat",
        "created": "Yaratilgan",
        "total_products": "Jami mahsulot",
        "total_stock": "Jami stok (kg)",
        "low": "Kam qolgan",
        "out": "Tugagan",
        "date": "Sana",
        "date_range": "Sana oralig‘i",
        "opening": "Boshlang‘ich (kg)",
        "closing": "Yakuniy (kg)",
        "stock_in": "Kirim (kg)",
        "stock_out": "Chiqim (kg)",
        "adjustment_in": "Musbat tuzatish",
        "adjustment_out": "Manfiy tuzatish",
        "net": "Sof o‘zgarish",
        "transactions": "Operatsiyalar",
        "affected": "Ta’sirlangan mahsulot",
        "normal": "Normal",
        "movement_type": "Harakat turi",
        "movement_all": "Barchasi",
        "movement_in": "Kirim",
        "movement_out": "Chiqim",
    },
}


def _decimal(value: Decimal | int | str) -> str:
    result = f"{Decimal(value):,.3f}"
    return result.rstrip("0").rstrip(".")


def _status(value: str, labels: dict[str, str]) -> str:
    return labels.get(value, value)


def product_report_document(
    items: list[ProductReportItem],
    summary: ProductReportSummary,
    filters: ProductReportFilters,
    language: ReportLanguage,
    generated_at: datetime,
) -> ReportDocument:
    labels = TRANSLATIONS[language]
    filter_parts = [
        value
        for value in (
            filters.search and f"search={filters.search}",
            filters.brand and f"brand={filters.brand}",
            filters.color and f"color={filters.color}",
            filters.lot_number and f"lot={filters.lot_number}",
            filters.stock_status.value != "all" and f"status={filters.stock_status.value}",
            filters.created_from and f"from={filters.created_from.isoformat()}",
            filters.created_to and f"to={filters.created_to.isoformat()}",
        )
        if value
    ]
    subtitle = f"{labels['generated']}: {generated_at:%Y-%m-%d %H:%M}"
    if filter_parts:
        subtitle += " | " + ", ".join(filter_parts)
    rows = [
        {
            "code": item.product_code,
            "name": item.name,
            "brand": item.brand or "—",
            "color": item.color or "—",
            "lot": item.lot_number,
            "current": _decimal(item.current_stock),
            "count": str(item.count) if item.count is not None else "-",
            "status": _status(item.stock_status, labels),
            "created": item.created_at.strftime("%Y-%m-%d"),
        }
        for item in items
    ]
    return ReportDocument(
        title=f"ALFATEKS — {labels['product_report']}",
        subtitle=subtitle,
        columns=[
            ExportColumn("code", labels["code"], 15),
            ExportColumn("name", labels["name"], 28),
            ExportColumn("brand", labels["brand"], 18),
            ExportColumn("color", labels["color"], 14),
            ExportColumn("lot", labels["lot"], 18),
            ExportColumn("current", labels["current"], 14),
            ExportColumn("count", labels["count"], 13),
            ExportColumn("status", labels["status"], 14),
            ExportColumn("created", labels["created"], 14),
        ],
        rows=rows,
        summary=[
            (labels["total_products"], summary.total_products),
            (labels["total_stock"], _decimal(summary.total_current_stock)),
            (labels["low"], summary.low_stock_products),
            (labels["out"], summary.out_of_stock_products),
        ],
        filename=f"alfateks-products-{generated_at:%Y%m%d-%H%M}",
    )


def daily_report_document(
    report: DailyStockReportResponse, language: ReportLanguage
) -> ReportDocument:
    labels = TRANSLATIONS[language]
    movement_label = labels[f"movement_{report.movement_type.value}"]
    date_value = (
        report.report_date_from.isoformat()
        if report.report_date_from == report.report_date_to
        else f"{report.report_date_from.isoformat()} – {report.report_date_to.isoformat()}"
    )
    rows = [
        {
            "code": item.product_code,
            "name": item.product_name,
            "opening": _decimal(item.opening_stock),
            "stock_in": _decimal(item.stock_in),
            "stock_out": _decimal(item.stock_out),
            "adjustment_in": _decimal(item.adjustment_in),
            "adjustment_out": _decimal(item.adjustment_out),
            "closing": _decimal(item.closing_stock),
            "net": _decimal(item.net_change),
            "transactions": item.transaction_count,
        }
        for item in report.products
    ]
    return ReportDocument(
        title=f"ALFATEKS — {labels['daily_report']}",
        subtitle=(
            f"{labels['date_range']}: {date_value} | "
            f"{labels['movement_type']}: {movement_label} | "
            f"{labels['generated']}: {report.generated_at:%Y-%m-%d %H:%M} | "
            f"TZ: {report.timezone}"
        ),
        columns=[
            ExportColumn("code", labels["code"], 15),
            ExportColumn("name", labels["name"], 25),
            ExportColumn("opening", labels["opening"], 14),
            ExportColumn("stock_in", labels["stock_in"], 14),
            ExportColumn("stock_out", labels["stock_out"], 14),
            ExportColumn("adjustment_in", labels["adjustment_in"], 16),
            ExportColumn("adjustment_out", labels["adjustment_out"], 16),
            ExportColumn("closing", labels["closing"], 14),
            ExportColumn("net", labels["net"], 14),
            ExportColumn("transactions", labels["transactions"], 12),
        ],
        rows=rows,
        summary=[
            (labels["stock_in"], _decimal(report.summary.stock_in)),
            (labels["stock_out"], _decimal(report.summary.stock_out)),
            (labels["adjustment_in"], _decimal(report.summary.adjustment_in)),
            (labels["adjustment_out"], _decimal(report.summary.adjustment_out)),
            (labels["net"], _decimal(report.summary.net_change)),
            (labels["transactions"], report.summary.transaction_count),
            (labels["affected"], report.summary.affected_products),
        ],
        filename=(
            f"alfateks-daily-stock-{report.report_date_from:%Y%m%d}"
            if report.report_date_from == report.report_date_to
            else (f"alfateks-stock-{report.report_date_from:%Y%m%d}-{report.report_date_to:%Y%m%d}")
        ),
    )


def export_document(document: ReportDocument, report_format: ReportFormat) -> ExportArtifact:
    normalized = report_format.normalized
    if normalized == ReportFormat.PDF:
        return ExportArtifact(_pdf(document), "application/pdf", f"{document.filename}.pdf")
    if normalized == ReportFormat.PNG:
        return ExportArtifact(_png(document), "image/png", f"{document.filename}.png")
    return ExportArtifact(
        _xlsx(document),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        f"{document.filename}.xlsx",
    )


def _register_pdf_fonts() -> tuple[str, str]:
    regular_name = "AlfateksVera"
    bold_name = "AlfateksVeraBold"
    if regular_name not in pdfmetrics.getRegisteredFontNames():
        fonts = Path(reportlab.__file__).resolve().parent / "fonts"
        pdfmetrics.registerFont(TTFont(regular_name, str(fonts / "Vera.ttf")))
        pdfmetrics.registerFont(TTFont(bold_name, str(fonts / "VeraBd.ttf")))
    return regular_name, bold_name


def _pdf(document: ReportDocument) -> bytes:
    regular, bold = _register_pdf_fonts()
    output = BytesIO()
    pdf = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=document.title,
        author="Alfateks Warehouse",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontName=bold,
        fontSize=17,
        textColor=colors.HexColor("#16332A"),
        alignment=TA_CENTER,
    )
    body_style = ParagraphStyle(
        "ReportBody",
        parent=styles["BodyText"],
        fontName=regular,
        fontSize=7.3,
        leading=9,
    )
    header_style = ParagraphStyle(
        "ReportHeader",
        parent=body_style,
        fontName=bold,
        textColor=colors.white,
    )
    story = [
        Paragraph(escape(document.title), title_style),
        Paragraph(escape(document.subtitle), body_style),
        Spacer(1, 5 * mm),
        Paragraph(
            " &nbsp; | &nbsp; ".join(
                f"<b>{escape(str(label))}:</b> {escape(str(value))}"
                for label, value in document.summary
            ),
            body_style,
        ),
        Spacer(1, 4 * mm),
    ]
    table_rows = [[Paragraph(escape(column.title), header_style) for column in document.columns]]
    for row in document.rows:
        table_rows.append(
            [
                Paragraph(escape(str(row.get(column.key, ""))), body_style)
                for column in document.columns
            ]
        )
    available_width = landscape(A4)[0] - 20 * mm
    column_widths = _pdf_column_widths(document, regular, bold, available_width)
    table = Table(table_rows, colWidths=column_widths, repeatRows=1, hAlign="CENTER")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#176B54")),
                ("FONTNAME", (0, 0), (-1, 0), bold),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(table)

    def page_number(canvas, doc) -> None:
        canvas.saveState()
        canvas.setFont(regular, 7)
        canvas.drawRightString(landscape(A4)[0] - 10 * mm, 6 * mm, f"{doc.page}")
        canvas.restoreState()

    pdf.build(story, onFirstPage=page_number, onLaterPages=page_number)
    return output.getvalue()


def _excel_safe(value: Any) -> Any:
    if isinstance(value, str) and value != "-" and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _visible_character_width(value: Any) -> int:
    """Approximate an Excel character width without stretching short columns."""
    return sum(2 if ord(character) > 0xFF else 1 for character in str(value))


def _excel_column_widths(document: ReportDocument) -> list[float]:
    widths: list[float] = []
    for column in document.columns:
        header_width = _visible_character_width(column.title)
        content_width = max(
            (_visible_character_width(row.get(column.key, "")) for row in document.rows),
            default=0,
        )
        maximum_width = column.width * 1.2
        widths.append(float(min(maximum_width, max(8, header_width + 3, content_width + 3))))
    return widths


def _xlsx(document: ReportDocument) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Alfateks Report"
    last_column = get_column_letter(len(document.columns))
    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = document.title
    sheet["A1"].font = Font(size=16, bold=True, color="176B54")
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells(f"A2:{last_column}2")
    sheet["A2"] = document.subtitle
    sheet["A2"].alignment = Alignment(horizontal="center")
    sheet.merge_cells(f"A4:{last_column}4")
    sheet["A4"] = " | ".join(f"{label}: {value}" for label, value in document.summary)
    sheet["A4"].font = Font(bold=True, color="334155")

    header_row = 6
    column_widths = _excel_column_widths(document)
    sheet.row_dimensions[header_row].height = 28
    for index, (column, column_width) in enumerate(
        zip(document.columns, column_widths, strict=True), start=1
    ):
        cell = sheet.cell(header_row, index, column.title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="176B54")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = column_width
    for row_index, row in enumerate(document.rows, start=header_row + 1):
        for column_index, column in enumerate(document.columns, start=1):
            cell = sheet.cell(row_index, column_index, _excel_safe(row.get(column.key, "")))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")
    last_row = max(header_row, header_row + len(document.rows))
    sheet.auto_filter.ref = f"A{header_row}:{last_column}{last_row}"
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.print_title_rows = f"{header_row}:{header_row}"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _load_png_font(size: int, *, bold: bool = False):
    bundled_fonts = Path(reportlab.__file__).resolve().parent / "fonts"
    candidates = (
        str(bundled_fonts / ("VeraBd.ttf" if bold else "Vera.ttf")),
        (
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
            if bold
            else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
        ),
        (
            "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf"
            if bold
            else "/usr/share/fonts/dejavu/DejaVuSans.ttf"
        ),
        "DejaVuSans-Bold.ttf" if bold else "DejaVuSans.ttf",
    )
    for candidate in candidates:
        try:
            return ImageFont.truetype(candidate, size)
        except OSError:
            continue
    return ImageFont.load_default()


def _fit_widths(preferred: list[float], minimum: list[float], available: float) -> list[float]:
    if sum(preferred) <= available:
        return preferred

    minimum_total = sum(minimum)
    if minimum_total >= available:
        scale = available / minimum_total
        return [width * scale for width in minimum]

    available_slack = available - minimum_total
    preferred_slack = sum(
        preferred_width - minimum_width
        for preferred_width, minimum_width in zip(preferred, minimum, strict=True)
    )
    return [
        minimum_width + (preferred_width - minimum_width) * available_slack / preferred_slack
        for preferred_width, minimum_width in zip(preferred, minimum, strict=True)
    ]


def _expand_widths(widths: list[float], target: float, weights: list[float]) -> list[float]:
    current = sum(widths)
    if current >= target:
        return widths

    weight_total = sum(weights)
    extra = target - current
    return [
        width + extra * weight / weight_total for width, weight in zip(widths, weights, strict=True)
    ]


def _pdf_column_widths(
    document: ReportDocument,
    regular_font: str,
    bold_font: str,
    available_width: float,
) -> list[float]:
    horizontal_padding = 8.0
    preferred: list[float] = []
    minimum: list[float] = []
    for column in document.columns:
        header_width = pdfmetrics.stringWidth(column.title, bold_font, 7.3)
        header_token_width = max(
            (pdfmetrics.stringWidth(token, bold_font, 7.3) for token in column.title.split()),
            default=0.0,
        )
        content_width = max(
            (
                pdfmetrics.stringWidth(str(row.get(column.key, "")), regular_font, 7.3)
                for row in document.rows
            ),
            default=0.0,
        )
        content_token_width = max(
            (
                pdfmetrics.stringWidth(token, regular_font, 7.3)
                for row in document.rows
                for token in str(row.get(column.key, "")).split()
            ),
            default=0.0,
        )
        preferred_width = max(30.0, header_width + horizontal_padding, content_width + 8.0)
        preferred.append(preferred_width)
        minimum.append(
            min(
                preferred_width,
                max(
                    30.0,
                    min(70.0, max(header_token_width + 8.0, content_token_width + 8.0)),
                ),
            )
        )

    fitted = _fit_widths(preferred, minimum, available_width)
    minimum_table_width = available_width * TABLE_MIN_WIDTH_RATIO
    return _expand_widths(
        fitted,
        minimum_table_width,
        [float(column.width) for column in document.columns],
    )


def _text_width(draw: ImageDraw.ImageDraw, value: Any, font: Any) -> float:
    return float(draw.textlength(str(value), font=font))


def _wrap_text(draw: ImageDraw.ImageDraw, value: Any, font: Any, max_width: float) -> list[str]:
    text = str(value)
    wrapped: list[str] = []
    for paragraph in text.splitlines() or [""]:
        remaining = paragraph.strip()
        if not remaining:
            wrapped.append("")
            continue

        while _text_width(draw, remaining, font) > max_width:
            low = 1
            high = len(remaining)
            while low < high:
                middle = (low + high + 1) // 2
                if _text_width(draw, remaining[:middle], font) <= max_width:
                    low = middle
                else:
                    high = middle - 1

            split_at = remaining.rfind(" ", 0, low + 1)
            if split_at > 0:
                wrapped.append(remaining[:split_at].rstrip())
                remaining = remaining[split_at + 1 :].lstrip()
            else:
                wrapped.append(remaining[:low])
                remaining = remaining[low:]
        wrapped.append(remaining)
    return wrapped


def _png_column_widths(
    draw: ImageDraw.ImageDraw,
    document: ReportDocument,
    header_font: Any,
    body_font: Any,
    available_width: int,
) -> list[int]:
    preferred: list[float] = []
    minimum: list[float] = []
    padding = PNG_CELL_HORIZONTAL_PADDING * 2
    for column in document.columns:
        header_width = _text_width(draw, column.title, header_font)
        header_token_width = max(
            (_text_width(draw, token, header_font) for token in column.title.split()),
            default=0.0,
        )
        content_width = max(
            (_text_width(draw, row.get(column.key, ""), body_font) for row in document.rows),
            default=0.0,
        )
        content_token_width = max(
            (
                _text_width(draw, token, body_font)
                for row in document.rows
                for token in str(row.get(column.key, "")).split()
            ),
            default=0.0,
        )
        preferred_width = max(PNG_MIN_COLUMN_WIDTH, header_width + padding, content_width + padding)
        preferred.append(preferred_width)
        minimum.append(
            min(
                preferred_width,
                max(
                    PNG_MIN_COLUMN_WIDTH,
                    min(
                        PNG_MAX_UNBROKEN_WIDTH,
                        max(header_token_width + padding, content_token_width + padding),
                    ),
                ),
            )
        )

    fitted = _fit_widths(preferred, minimum, float(available_width))
    widths = [max(1, round(width)) for width in fitted]
    overflow = sum(widths) - available_width
    for index in range(len(widths) - 1, -1, -1):
        if overflow <= 0:
            break
        reduction = min(overflow, widths[index] - 1)
        widths[index] -= reduction
        overflow -= reduction
    return widths


def _draw_text_lines(
    draw: ImageDraw.ImageDraw,
    xy: tuple[float, float],
    height: float,
    lines: list[str],
    *,
    font: Any,
    fill: str,
    line_height: int,
) -> None:
    block_height = len(lines) * line_height
    line_top = xy[1] + (height - block_height) / 2
    for index, line in enumerate(lines):
        bounds = draw.textbbox((0, 0), line, font=font)
        text_height = bounds[3] - bounds[1]
        y = line_top + index * line_height + (line_height - text_height) / 2 - bounds[1]
        draw.text((xy[0], y), line, font=font, fill=fill)


def _png(document: ReportDocument) -> bytes:
    measurement_image = Image.new("RGB", (1, 1), "white")
    draw = ImageDraw.Draw(measurement_image)
    title_font = _load_png_font(60, bold=True)
    subtitle_font = _load_png_font(28)
    header_font = _load_png_font(30, bold=True)
    body_font = _load_png_font(30)

    widths = _png_column_widths(draw, document, header_font, body_font, PNG_MAX_WIDTH)
    table_width = sum(widths)
    table_left = 0
    table_right = table_width - 1

    header_text_width = table_width - PNG_HEADER_HORIZONTAL_PADDING * 2
    title_lines = _wrap_text(draw, document.title, title_font, header_text_width)
    subtitle_lines = _wrap_text(draw, document.subtitle, subtitle_font, header_text_width)
    title_block_height = len(title_lines) * PNG_TITLE_LINE_HEIGHT
    subtitle_block_height = len(subtitle_lines) * PNG_SUBTITLE_LINE_HEIGHT
    report_header_height = max(
        PNG_HEADER_MIN_HEIGHT,
        PNG_HEADER_VERTICAL_PADDING * 2
        + title_block_height
        + PNG_HEADER_TEXT_GAP
        + subtitle_block_height,
    )

    summary = "  |  ".join(f"{label}: {value}" for label, value in document.summary)
    summary_lines = _wrap_text(draw, summary, body_font, table_width)
    summary_height = (
        PNG_SUMMARY_TOP_PADDING
        + len(summary_lines) * PNG_TABLE_BODY_LINE_HEIGHT
        + PNG_SUMMARY_BOTTOM_PADDING
    )

    header_cells = [
        _wrap_text(
            draw,
            column.title,
            header_font,
            width - PNG_CELL_HORIZONTAL_PADDING * 2,
        )
        for column, width in zip(document.columns, widths, strict=True)
    ]
    table_header_height = max(
        PNG_TABLE_HEADER_MIN_HEIGHT,
        max((len(lines) for lines in header_cells), default=1) * PNG_TABLE_HEADER_LINE_HEIGHT
        + PNG_CELL_VERTICAL_PADDING * 2,
    )

    source_rows = document.rows or [{}]
    row_cells: list[list[list[str]]] = []
    row_heights: list[int] = []
    for row in source_rows:
        cells = [
            _wrap_text(
                draw,
                row.get(column.key, ""),
                body_font,
                width - PNG_CELL_HORIZONTAL_PADDING * 2,
            )
            for column, width in zip(document.columns, widths, strict=True)
        ]
        row_cells.append(cells)
        row_heights.append(
            max(
                PNG_TABLE_ROW_MIN_HEIGHT,
                max((len(lines) for lines in cells), default=1) * PNG_TABLE_BODY_LINE_HEIGHT
                + PNG_CELL_VERTICAL_PADDING * 2,
            )
        )

    table_top = report_header_height + summary_height
    table_bottom = table_top + table_header_height + sum(row_heights)
    image = Image.new("RGB", (table_width, table_bottom), "white")
    draw = ImageDraw.Draw(image)

    report_header_top = 0
    report_header_bottom = report_header_top + report_header_height
    draw.rectangle(
        (table_left, report_header_top, table_right, report_header_bottom - 1),
        fill="#16332A",
    )
    draw.rectangle(
        (
            table_left,
            report_header_bottom - PNG_ACCENT_HEIGHT,
            table_right,
            report_header_bottom - 1,
        ),
        fill="#2F9E7D",
    )
    title_top = report_header_top + PNG_HEADER_VERTICAL_PADDING
    _draw_text_lines(
        draw,
        (table_left + PNG_HEADER_HORIZONTAL_PADDING, title_top),
        title_block_height,
        title_lines,
        font=title_font,
        fill="white",
        line_height=PNG_TITLE_LINE_HEIGHT,
    )
    _draw_text_lines(
        draw,
        (
            table_left + PNG_HEADER_HORIZONTAL_PADDING,
            title_top + title_block_height + PNG_HEADER_TEXT_GAP,
        ),
        subtitle_block_height,
        subtitle_lines,
        font=subtitle_font,
        fill="#D8E7E2",
        line_height=PNG_SUBTITLE_LINE_HEIGHT,
    )

    _draw_text_lines(
        draw,
        (table_left, report_header_bottom + PNG_SUMMARY_TOP_PADDING),
        len(summary_lines) * PNG_TABLE_BODY_LINE_HEIGHT,
        summary_lines,
        font=body_font,
        fill="#475569",
        line_height=PNG_TABLE_BODY_LINE_HEIGHT,
    )

    draw.rectangle(
        (table_left, table_top, table_right, table_top + table_header_height - 1),
        fill="#176B54",
    )
    x = table_left
    for lines, width in zip(header_cells, widths, strict=True):
        _draw_text_lines(
            draw,
            (x + PNG_CELL_HORIZONTAL_PADDING, table_top),
            table_header_height,
            lines,
            font=header_font,
            fill="white",
            line_height=PNG_TABLE_HEADER_LINE_HEIGHT,
        )
        x += width

    y = table_top + table_header_height
    for row_index, (cells, row_height) in enumerate(zip(row_cells, row_heights, strict=True)):
        if row_index % 2:
            draw.rectangle(
                (table_left, y, table_right, y + row_height - 1),
                fill="#F1F5F9",
            )
        x = table_left
        for lines, width in zip(cells, widths, strict=True):
            _draw_text_lines(
                draw,
                (x + PNG_CELL_HORIZONTAL_PADDING, y),
                row_height,
                lines,
                font=body_font,
                fill="#172033",
                line_height=PNG_TABLE_BODY_LINE_HEIGHT,
            )
            x += width
        draw.line(
            (table_left, y + row_height - 1, table_right, y + row_height - 1),
            fill="#CBD5E1",
            width=PNG_GRID_LINE_WIDTH,
        )
        y += row_height

    x = table_left
    for width in widths[:-1]:
        x += width
        draw.line(
            (x, table_top, x, table_bottom - 1),
            fill="#CBD5E1",
            width=PNG_GRID_LINE_WIDTH,
        )
    draw.rectangle(
        (table_left, table_top, table_right, table_bottom - 1),
        outline="#B8C5D1",
        width=PNG_GRID_LINE_WIDTH,
    )
    output = BytesIO()
    image.save(
        output,
        format="PNG",
        optimize=True,
        dpi=(PNG_DPI, PNG_DPI),
    )
    return output.getvalue()
