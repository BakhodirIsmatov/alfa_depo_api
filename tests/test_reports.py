from dataclasses import replace
from datetime import UTC, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from zoneinfo import ZoneInfo

import httpx
import pytest
from openpyxl import load_workbook
from PIL import Image, ImageDraw
from sqlalchemy import update

from app.core.database import AsyncSessionFactory
from app.models.product import Product
from app.models.stock_transaction import StockTransaction, StockTransactionType
from app.services import report_export


@pytest.mark.asyncio
async def test_product_report_filters_summary_and_options(
    client: httpx.AsyncClient, auth_headers: dict[str, str], create_product
) -> None:
    await create_product(
        product_code="ALF-REPORT-H",
        name="Healthy Navy",
        lot_number="REPORT-H",
        brand="Report Brand A",
        color="Navy",
        initial_stock="50",
        minimum_stock="10",
        count=24,
    )
    await create_product(
        product_code="ALF-REPORT-L",
        name="Low Navy",
        lot_number="REPORT-L",
        brand="Report Brand A",
        color="Navy",
        initial_stock="5",
        minimum_stock="10",
    )
    await create_product(
        product_code="ALF-REPORT-E",
        name="Empty Black",
        lot_number="REPORT-O",
        brand="Report Brand B",
        color="Black",
        initial_stock="0",
        minimum_stock="10",
    )
    async with AsyncSessionFactory() as session:
        await session.execute(
            update(Product)
            .where(Product.product_code.in_(["ALF-REPORT-H", "ALF-REPORT-L"]))
            .values(color="Navy")
        )
        await session.execute(
            update(Product).where(Product.product_code == "ALF-REPORT-E").values(color="Black")
        )
        await session.commit()

    all_report = await client.get(
        "/api/v1/reports/products?page_size=2&sort_by=name&sort_order=asc",
        headers=auth_headers,
    )
    assert all_report.status_code == 200
    data = all_report.json()["data"]
    assert data["pagination"] == {
        "page": 1,
        "page_size": 2,
        "total": 3,
        "pages": 2,
    }
    assert data["summary"]["total_products"] == 3
    assert Decimal(data["summary"]["total_current_stock"]) == Decimal("55")
    assert data["summary"]["low_stock_products"] == 2
    assert data["summary"]["out_of_stock_products"] == 1
    counts_by_code = {item["product_code"]: item["count"] for item in data["items"]}
    assert counts_by_code == {"ALF-REPORT-E": None, "ALF-REPORT-H": 24}

    low_report = await client.get(
        "/api/v1/reports/products?brand=Report%20Brand%20A&stock_status=low",
        headers=auth_headers,
    )
    assert low_report.status_code == 200
    low_data = low_report.json()["data"]
    assert [item["name"] for item in low_data["items"]] == ["Low Navy"]
    assert low_data["items"][0]["stock_status"] == "low"

    options = await client.get("/api/v1/reports/products/filter-options", headers=auth_headers)
    assert options.status_code == 200
    assert options.json()["data"] == {
        "brands": ["Report Brand A", "Report Brand B"],
        "colors": ["Black", "Navy"],
    }


@pytest.mark.asyncio
async def test_product_report_rejects_invalid_ranges(
    client: httpx.AsyncClient, auth_headers: dict[str, str]
) -> None:
    response = await client.get(
        "/api/v1/reports/products?minimum_current_stock=20&maximum_current_stock=10",
        headers=auth_headers,
    )
    assert response.status_code == 422
    assert response.json()["error"]["code"] == "INVALID_REPORT_FILTERS"


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("report_format", "content_type", "signature"),
    [
        ("pdf", "application/pdf", b"%PDF"),
        ("png", "image/png", b"\x89PNG"),
        (
            "xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            b"PK",
        ),
        (
            "xls",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            b"PK",
        ),
    ],
)
async def test_product_report_exports_real_files(
    report_format: str,
    content_type: str,
    signature: bytes,
    client: httpx.AsyncClient,
    auth_headers: dict[str, str],
    create_product,
) -> None:
    await create_product(name="Export Product", lot_number="EXPORT-1")
    response = await client.get(
        f"/api/v1/reports/products/export?format={report_format}&language=uz",
        headers=auth_headers,
    )
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(content_type)
    assert response.headers["content-disposition"].endswith(
        f'.{"xlsx" if report_format in {"xls", "xlsx"} else report_format}"'
    )
    assert response.headers["x-report-row-count"] == "1"
    assert response.content.startswith(signature)


@pytest.mark.asyncio
async def test_product_report_xlsx_renders_nullable_count_as_dash(
    client: httpx.AsyncClient,
    auth_headers: dict[str, str],
    create_product,
) -> None:
    await create_product(
        product_code="ALF-COUNT-NULL",
        name="Count Null",
        lot_number="COUNT-NULL",
    )
    response = await client.get(
        "/api/v1/reports/products/export?format=xlsx&language=en",
        headers=auth_headers,
    )
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[6]]
    count_column = headers.index("Product count") + 1
    assert sheet.cell(7, count_column).value == "-"


@pytest.mark.asyncio
async def test_png_export_limit_is_enforced(
    client: httpx.AsyncClient,
    auth_headers: dict[str, str],
    create_product,
    monkeypatch,
) -> None:
    from app.core.config import get_settings

    await create_product(product_code="ALF-PNG-1", name="PNG One", lot_number="PNG-1")
    await create_product(product_code="ALF-PNG-2", name="PNG Two", lot_number="PNG-2")
    monkeypatch.setattr(get_settings(), "report_png_max_rows", 1)
    response = await client.get("/api/v1/reports/products/export?format=png", headers=auth_headers)
    assert response.status_code == 422
    assert response.json()["error"]["code"] == "REPORT_TOO_LARGE"


@pytest.mark.asyncio
async def test_png_export_uses_dynamic_300_dpi_canvas_with_a4_max_width(
    client: httpx.AsyncClient,
    auth_headers: dict[str, str],
    create_product,
) -> None:
    await create_product(name="PNG A4", lot_number="PNG-A4")
    response = await client.get("/api/v1/reports/products/export?format=png", headers=auth_headers)
    assert response.status_code == 200
    image = Image.open(BytesIO(response.content))
    assert image.width > image.height
    assert image.width < report_export.PNG_MAX_WIDTH
    assert image.info["dpi"] == pytest.approx((300, 300), abs=0.1)
    bottom_row = image.convert("RGB").crop((0, image.height - 1, image.width, image.height))
    assert bottom_row.getextrema() != ((255, 255), (255, 255), (255, 255))


@pytest.mark.asyncio
async def test_export_layout_compacts_columns_and_balances_png_header(
    client: httpx.AsyncClient,
    auth_headers: dict[str, str],
    create_product,
) -> None:
    await create_product(
        product_code="P1",
        name="Atlas",
        lot_number="L1",
        brand="A",
    )

    png_response = await client.get(
        "/api/v1/reports/products/export?format=png&language=uz",
        headers=auth_headers,
    )
    assert png_response.status_code == 200
    image = Image.open(BytesIO(png_response.content)).convert("RGB")
    report_header_height = max(
        report_export.PNG_HEADER_MIN_HEIGHT,
        report_export.PNG_HEADER_VERTICAL_PADDING * 2
        + report_export.PNG_TITLE_LINE_HEIGHT
        + report_export.PNG_HEADER_TEXT_GAP
        + report_export.PNG_SUBTITLE_LINE_HEIGHT,
    )
    accent_y = report_header_height - 2
    table_y = (
        report_header_height
        + report_export.PNG_SUMMARY_TOP_PADDING
        + report_export.PNG_TABLE_BODY_LINE_HEIGHT
        + report_export.PNG_SUMMARY_BOTTOM_PADDING
        + 1
    )
    accent_x = [x for x in range(image.width) if image.getpixel((x, accent_y)) == (47, 158, 125)]
    table_x = [x for x in range(image.width) if image.getpixel((x, table_y)) != (255, 255, 255)]
    assert (min(accent_x), max(accent_x)) == (0, image.width - 1)
    assert (min(table_x), max(table_x)) == (0, image.width - 1)
    assert image.width < report_export.PNG_MAX_WIDTH

    xlsx_response = await client.get(
        "/api/v1/reports/products/export?format=xlsx&language=uz",
        headers=auth_headers,
    )
    assert xlsx_response.status_code == 200
    workbook = load_workbook(BytesIO(xlsx_response.content), read_only=False)
    sheet = workbook.active
    assert 16 <= sheet.column_dimensions["B"].width < 28
    assert sheet.row_dimensions[6].height == 28


def test_png_text_wrap_preserves_long_values() -> None:
    image = Image.new("RGB", (200, 100), "white")
    draw = ImageDraw.Draw(image)
    font = report_export._load_png_font(13)
    value = "Uzun mahsulot nomi barcha so‘zlari bilan saqlanishi kerak"

    lines = report_export._wrap_text(draw, value, font, 110)

    assert len(lines) > 1
    assert " ".join(lines) == value
    assert all("…" not in line for line in lines)


def test_png_font_loader_keeps_requested_high_resolution_size() -> None:
    regular = report_export._load_png_font(30)
    bold = report_export._load_png_font(30, bold=True)

    assert getattr(regular, "size", None) == 30
    assert getattr(bold, "size", None) == 30


def test_png_canvas_tracks_dynamic_table_width_and_height() -> None:
    short_document = report_export.ReportDocument(
        title="ALFATEKS",
        subtitle="2026-09-04",
        columns=[
            report_export.ExportColumn("code", "Code", 12),
            report_export.ExportColumn("name", "Name", 28),
        ],
        rows=[{"code": "P1", "name": "Atlas"}],
        summary=[("Products", 1)],
        filename="short",
    )
    expanded_document = replace(
        short_document,
        rows=[
            {
                "code": f"ALF-{index:04d}-" + "X" * 72,
                "name": "Professional textile product with a fully preserved long name",
            }
            for index in range(8)
        ],
    )

    short_image = Image.open(BytesIO(report_export._png(short_document)))
    expanded_image = Image.open(BytesIO(report_export._png(expanded_document)))

    assert expanded_image.width > short_image.width
    assert expanded_image.height > short_image.height
    assert expanded_image.width <= report_export.PNG_MAX_WIDTH


def test_pdf_widths_balance_short_content_and_expand_for_long_values() -> None:
    regular, bold = report_export._register_pdf_fonts()
    short_document = report_export.ReportDocument(
        title="Report",
        subtitle="Generated",
        columns=[
            report_export.ExportColumn("code", "Code", 12),
            report_export.ExportColumn("name", "Name", 28),
        ],
        rows=[{"code": "P1", "name": "Atlas"}],
        summary=[],
        filename="report",
    )
    available_width = 700.0

    short_widths = report_export._pdf_column_widths(
        short_document,
        regular,
        bold,
        available_width,
    )
    long_document = replace(
        short_document,
        rows=[{"code": "P1", "name": "Professional textile product " * 10}],
    )
    long_widths = report_export._pdf_column_widths(
        long_document,
        regular,
        bold,
        available_width,
    )

    assert sum(short_widths) == pytest.approx(available_width * report_export.TABLE_MIN_WIDTH_RATIO)
    assert short_widths[1] > short_widths[0]
    assert long_widths[1] > short_widths[1]
    assert sum(long_widths) <= available_width


@pytest.mark.asyncio
async def test_daily_dashboard_report_calculates_in_out_and_adjustments(
    client: httpx.AsyncClient, auth_headers: dict[str, str], create_product
) -> None:
    product = await create_product(name="Daily Movement", lot_number="DAILY-1", initial_stock="0")
    stock_in = await client.post(
        f"/api/v1/products/{product['id']}/stock/in",
        json={"quantity": "10", "count": 10, "note": "daily receive"},
        headers=auth_headers,
    )
    stock_out = await client.post(
        f"/api/v1/products/{product['id']}/stock/out",
        json={"quantity": "3", "count": 3, "note": "daily issue"},
        headers=auth_headers,
    )
    adjustment = await client.post(
        f"/api/v1/products/{product['id']}/stock/adjust",
        json={"quantity": "12", "count": 2, "note": "count"},
        headers=auth_headers,
    )
    assert stock_in.status_code == stock_out.status_code == adjustment.status_code == 201

    response = await client.get("/api/v1/dashboard/daily", headers=auth_headers)
    assert response.status_code == 200
    data = response.json()["data"]
    summary = data["summary"]
    assert Decimal(summary["stock_in"]) == Decimal("10")
    assert Decimal(summary["stock_out"]) == Decimal("3")
    assert Decimal(summary["adjustment_in"]) == Decimal("12")
    assert Decimal(summary["adjustment_out"]) == Decimal("0")
    assert Decimal(summary["net_change"]) == Decimal("19")
    assert summary["transaction_count"] == 3
    assert summary["affected_products"] == 1
    movement = data["products"][0]
    assert Decimal(movement["opening_stock"]) == Decimal("0")
    assert Decimal(movement["closing_stock"]) == Decimal("19")

    exported = await client.get(
        f"/api/v1/dashboard/daily/export?date={data['report_date']}&format=xlsx",
        headers=auth_headers,
    )
    assert exported.status_code == 200
    assert exported.content.startswith(b"PK")


@pytest.mark.asyncio
async def test_daily_report_aggregates_an_inclusive_date_range(
    client: httpx.AsyncClient, auth_headers: dict[str, str], create_product
) -> None:
    product = await create_product(
        product_code="ALF-RANGE-1",
        name="Range Movement",
        lot_number="RANGE-1",
        initial_stock="0",
    )
    stock_in = await client.post(
        f"/api/v1/products/{product['id']}/stock/in",
        json={"quantity": "10", "count": 10, "note": "range in"},
        headers=auth_headers,
    )
    stock_out = await client.post(
        f"/api/v1/products/{product['id']}/stock/out",
        json={"quantity": "3", "count": 3, "note": "range out"},
        headers=auth_headers,
    )
    assert stock_in.status_code == stock_out.status_code == 201

    reporting_now = datetime.now(ZoneInfo("Europe/Istanbul"))
    date_from = reporting_now.date() - timedelta(days=2)
    date_to = reporting_now.date()
    async with AsyncSessionFactory() as session:
        await session.execute(
            update(StockTransaction)
            .where(
                StockTransaction.product_id == product["id"],
                StockTransaction.note == "range in",
            )
            .values(created_at=reporting_now.replace(hour=12) - timedelta(days=2))
        )
        await session.commit()

    response = await client.get(
        f"/api/v1/dashboard/daily?date_from={date_from}&date_to={date_to}",
        headers=auth_headers,
    )
    assert response.status_code == 200
    data = response.json()["data"]
    assert data["report_date_from"] == date_from.isoformat()
    assert data["report_date_to"] == date_to.isoformat()
    assert Decimal(data["summary"]["stock_in"]) == Decimal("10")
    assert Decimal(data["summary"]["stock_out"]) == Decimal("3")
    assert Decimal(data["products"][0]["opening_stock"]) == Decimal("0")
    assert Decimal(data["products"][0]["closing_stock"]) == Decimal("7")

    exported = await client.get(
        (f"/api/v1/dashboard/daily/export?date_from={date_from}&date_to={date_to}&format=xlsx"),
        headers=auth_headers,
    )
    assert exported.status_code == 200
    workbook = load_workbook(BytesIO(exported.content), read_only=True)
    assert f"Tarih aralığı: {date_from} – {date_to}" in workbook.active["A2"].value


@pytest.mark.asyncio
async def test_daily_report_rejects_reversed_date_range(
    client: httpx.AsyncClient, auth_headers: dict[str, str]
) -> None:
    response = await client.get(
        "/api/v1/dashboard/daily?date_from=2026-09-04&date_to=2026-09-01",
        headers=auth_headers,
    )
    assert response.status_code == 422
    assert response.json()["error"]["code"] == "INVALID_REPORT_FILTERS"


@pytest.mark.asyncio
async def test_daily_report_direction_filter_is_shared_by_preview_and_export(
    client: httpx.AsyncClient, auth_headers: dict[str, str], create_product
) -> None:
    inbound = await create_product(
        product_code="ALF-DAILY-IN",
        name="Daily In",
        lot_number="DAILY-IN",
        initial_stock="0",
    )
    outbound = await create_product(
        product_code="ALF-DAILY-OUT",
        name="Daily Out",
        lot_number="DAILY-OUT",
        initial_stock="10",
    )
    adjustment_in = await create_product(
        product_code="ALF-DAILY-ADJ-IN",
        name="Daily Adjustment In",
        lot_number="DAILY-ADJ-IN",
        initial_stock="0",
    )
    adjustment_out = await create_product(
        product_code="ALF-DAILY-ADJ-OUT",
        name="Daily Adjustment Out",
        lot_number="DAILY-ADJ-OUT",
        initial_stock="10",
        count=10,
    )
    async with AsyncSessionFactory() as session:
        await session.execute(
            update(StockTransaction)
            .where(
                StockTransaction.product_id.in_([outbound["id"], adjustment_out["id"]]),
                StockTransaction.transaction_type == StockTransactionType.IN,
            )
            .values(created_at=datetime.now(UTC) - timedelta(days=2))
        )
        await session.commit()
    assert (
        await client.post(
            f"/api/v1/products/{inbound['id']}/stock/in",
            json={"quantity": "5", "count": 5, "note": "direction in"},
            headers=auth_headers,
        )
    ).status_code == 201
    assert (
        await client.post(
            f"/api/v1/products/{adjustment_in['id']}/stock/adjust",
            json={"quantity": "2", "count": 2, "note": "positive direction"},
            headers=auth_headers,
        )
    ).status_code == 201
    assert (
        await client.post(
            f"/api/v1/products/{adjustment_out['id']}/stock/adjust",
            json={"quantity": "-4", "count": -4, "note": "negative direction"},
            headers=auth_headers,
        )
    ).status_code == 201
    assert (
        await client.post(
            f"/api/v1/products/{outbound['id']}/stock/out",
            json={"quantity": "4", "count": 4, "note": "direction out"},
            headers=auth_headers,
        )
    ).status_code == 201

    inbound_preview = await client.get(
        "/api/v1/dashboard/daily?movement_type=in", headers=auth_headers
    )
    assert inbound_preview.status_code == 200
    inbound_data = inbound_preview.json()["data"]
    assert inbound_data["movement_type"] == "in"
    assert [item["product_code"] for item in inbound_data["products"]] == [
        "ALF-DAILY-ADJ-IN",
        "ALF-DAILY-IN",
    ]
    assert Decimal(inbound_data["summary"]["stock_in"]) == Decimal("5")
    assert Decimal(inbound_data["summary"]["adjustment_in"]) == Decimal("2")
    assert Decimal(inbound_data["summary"]["stock_out"]) == Decimal("0")
    assert inbound_data["summary"]["transaction_count"] == 2

    outbound_preview = await client.get(
        "/api/v1/dashboard/daily?movement_type=out", headers=auth_headers
    )
    assert outbound_preview.status_code == 200
    outbound_data = outbound_preview.json()["data"]
    assert outbound_data["movement_type"] == "out"
    assert [item["product_code"] for item in outbound_data["products"]] == [
        "ALF-DAILY-ADJ-OUT",
        "ALF-DAILY-OUT",
    ]
    assert Decimal(outbound_data["summary"]["stock_in"]) == Decimal("0")
    assert Decimal(outbound_data["summary"]["stock_out"]) == Decimal("4")
    assert Decimal(outbound_data["summary"]["adjustment_out"]) == Decimal("4")
    assert Decimal(outbound_data["summary"]["net_change"]) == Decimal("-8")

    exported = await client.get(
        "/api/v1/dashboard/daily/export?movement_type=in&format=xlsx",
        headers=auth_headers,
    )
    assert exported.status_code == 200
    assert exported.content.startswith(b"PK")
    assert exported.headers["x-report-row-count"] == "2"
    workbook = load_workbook(BytesIO(exported.content), read_only=True)
    assert "Hareket türü: Giriş" in workbook.active["A2"].value


@pytest.mark.asyncio
async def test_filtered_daily_png_export_uses_dynamic_high_resolution_canvas(
    client: httpx.AsyncClient, auth_headers: dict[str, str], create_product
) -> None:
    product = await create_product(
        product_code="ALF-DAILY-PNG",
        name="Daily PNG",
        lot_number="DAILY-PNG",
        initial_stock="0",
    )
    response = await client.post(
        f"/api/v1/products/{product['id']}/stock/in",
        json={"quantity": "5", "count": 5, "note": "A4 PNG"},
        headers=auth_headers,
    )
    assert response.status_code == 201

    exported = await client.get(
        "/api/v1/dashboard/daily/export?movement_type=in&format=png",
        headers=auth_headers,
    )
    assert exported.status_code == 200
    assert exported.headers["x-report-row-count"] == "1"
    image = Image.open(BytesIO(exported.content))
    assert image.width <= report_export.PNG_MAX_WIDTH
    assert image.height > 0
    assert image.info["dpi"] == pytest.approx((300, 300), abs=0.1)


@pytest.mark.asyncio
async def test_reports_require_authentication(client: httpx.AsyncClient) -> None:
    response = await client.get("/api/v1/reports/products")
    assert response.status_code == 401
